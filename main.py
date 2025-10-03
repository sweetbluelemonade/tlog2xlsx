import argparse
from pathlib import Path
from openpyxl import Workbook

from logtypes import LogProcessing as lp
def save_table(path, paged_table):
    wb = Workbook()

    for name, table in paged_table.items():
        ws = wb.create_sheet(title = name)

        for row in table:
            ws.append(row)

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(path)

def extend_paged_table(paged_table, paged_extension):
    for page_name in paged_table.keys():
        for page_name_ext in paged_extension.keys():
            if page_name_ext == page_name:
                paged_table[page_name].extend(paged_extension[page_name_ext])

def open_close(f):
    ff = f.open()
    ff.close()
    return ff

def dir_path(path_str):
    path = Path(path_str)
    if not path.is_dir():
        raise argparse.ArgumentTypeError(
            f'Not a valid directory: {path}')
    return path

def main():
    parser = argparse.ArgumentParser(
        prog='tlogtotables',
        description='description',
        epilog='text under help'
    )

    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('-f', '--file', help='Path to log file', 
                        action='store',
                        type=argparse.FileType('r'))
    group.add_argument('-d', '--dir', help='Path to log file', 
                        action='store',
                        type=dir_path)
    parser.add_argument('-o', '--output', 
                        help='Output table file', 
                        action='store', 
                        required=True)
    parser.add_argument('-t', '--type', choices=lp.TYPES, help='Log type', 
                        action='store', required=True)

    args = parser.parse_args()

    files = []

    if not args.file is None:
        files = [args.file]
        files[0].close()
    else:
        files = [open_close(f) for f in args.dir.iterdir() if f.is_file()]

    paged_table = lp.get_headers(args.type)

    for f in files:
        with open(f.name, 'r') as ff:
            data = ff.read()
            paged_table_extension = lp.get_rows(args.type, data)
            extend_paged_table(paged_table, paged_table_extension)

    save_table(args.output, paged_table)

            

if __name__ == '__main__':
    main()